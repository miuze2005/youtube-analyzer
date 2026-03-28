"""
YouTube Comment Cleaner & Sentiment Analyzer
=============================================
Input : CSV or Excel with columns: username, timestamp, content, language, likes
Output: Cleaned & enriched Excel with topic, keywords, sentiment label, sentiment score
"""

import os
import re
import json
import time
import argparse
import unicodedata
import pandas as pd
import anthropic
from pathlib import Path


# ─────────────────────────────────────────────
# STEP 1 – DATA CLEANING
# ─────────────────────────────────────────────

EMOJI_RE = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F300-\U0001F5FF"  # symbols & pictographs
    "\U0001F680-\U0001F9FF"  # transport, flags, etc.
    "\U00002600-\U000027BF"  # misc symbols
    "\U0001FA00-\U0001FFFF"
    "\U00002702-\U000027B0"
    "\uFE00-\uFE0F"          # variation selectors
    "\U00010000-\U0010FFFF"
    "]+",
    flags=re.UNICODE,
)

SPAM_PATTERNS = re.compile(
    r"(sub(?:scribe)?\s*(to\s*)?me|check\s*out\s*my|follow\s*me|"
    r"promo\s*code|discount|click\s*(the\s*)?link|t\.me/|bit\.ly/|"
    r"giveaway|free\s*(gift|money|coins|robux)|dm\s*me|whatsapp|"
    r"telegram|only\s*fans|onlyfans|visit\s*my|buy\s*now|"
    r"first\s*\d*\s*comment|❤️.*❤️.*❤️)",
    re.IGNORECASE,
)


def is_emoji_only(text: str) -> bool:
    stripped = EMOJI_RE.sub("", text).strip()
    # Also strip common punctuation and whitespace
    stripped = re.sub(r"[\s!?.,;:\-_~*#@&()\[\]\"']+", "", stripped)
    return len(stripped) == 0


def is_spam(text: str) -> bool:
    return bool(SPAM_PATTERNS.search(text))


def normalize(text: str) -> str:
    """Lower-case + collapse whitespace for duplicate detection."""
    return re.sub(r"\s+", " ", text.lower().strip())


def clean_dataframe(df: pd.DataFrame, min_len: int = 40, max_len: int = 500) -> pd.DataFrame:
    print(f"\n[CLEAN] Original rows : {len(df):,}")

    # Ensure correct column types
    df["content"] = df["content"].astype(str).str.strip()

    # 1. Drop blank / NaN
    df = df[df["content"].notna() & (df["content"] != "") & (df["content"].str.lower() != "nan")]
    print(f"[CLEAN] After blank drop  : {len(df):,}")

    # 2. Length filter (40–500 chars)
    df = df[df["content"].str.len().between(min_len, max_len)]
    print(f"[CLEAN] After length filter: {len(df):,}")

    # 3. Emoji-only comments
    df = df[~df["content"].apply(is_emoji_only)]
    print(f"[CLEAN] After emoji-only   : {len(df):,}")

    # 4. Spam / low-value promotional
    df = df[~df["content"].apply(is_spam)]
    print(f"[CLEAN] After spam filter  : {len(df):,}")

    # 5. Exact & near-duplicate removal (keep first occurrence)
    df["_norm"] = df["content"].apply(normalize)
    df = df.drop_duplicates(subset="_norm", keep="first")
    df = df.drop(columns=["_norm"])
    print(f"[CLEAN] After dedup        : {len(df):,}")

    df = df.reset_index(drop=True)
    return df


# ─────────────────────────────────────────────
# STEP 2 – AI ANALYSIS  (Anthropic Claude)
# ─────────────────────────────────────────────

SYSTEM_PROMPT = """You are a multilingual sentiment analysis and topic classification expert.
You will receive a batch of YouTube comments (possibly in different languages).
For EACH comment return a JSON array where every element has EXACTLY these keys:
  - "topics"   : array of up to 3 short topic strings (in English)
  - "keywords" : array of up to 5 keyword strings (in English)
  - "sentiment": one of "Positive", "Neutral", "Negative"
  - "score"    : float from -1.0 (most negative) to +1.0 (most positive)

Rules:
- Output ONLY the JSON array, no markdown, no explanation.
- Preserve the ORDER of comments (index 0 = first comment, etc.).
- Topics should be broad themes (e.g. "Music Quality", "Artist Performance", "Lyrics").
- Keywords should be the most meaningful content words.
- Score 0.0 means perfectly neutral; use decimals (e.g. 0.75, -0.4).
"""


def build_user_prompt(comments: list[str]) -> str:
    numbered = "\n".join(f"{i}. {c}" for i, c in enumerate(comments))
    return f"Analyze these {len(comments)} comments:\n\n{numbered}"


def parse_response(raw: str, expected: int) -> list[dict]:
    """Extract the JSON array from the model reply."""
    # Strip possible markdown fences
    raw = re.sub(r"```(?:json)?", "", raw).strip().rstrip("```").strip()
    data = json.loads(raw)
    if not isinstance(data, list):
        raise ValueError("Model did not return a JSON array.")
    if len(data) != expected:
        raise ValueError(f"Expected {expected} results, got {len(data)}.")
    return data


def analyze_batch(client: anthropic.Anthropic, comments: list[str], retries: int = 3) -> list[dict]:
    for attempt in range(1, retries + 1):
        try:
            msg = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": build_user_prompt(comments)}],
            )
            raw = msg.content[0].text
            return parse_response(raw, len(comments))
        except (json.JSONDecodeError, ValueError) as e:
            print(f"  [WARN] Parse error on attempt {attempt}: {e}")
            if attempt == retries:
                # Return neutral placeholders so the pipeline doesn't crash
                return [
                    {"topics": ["Unknown"], "keywords": [], "sentiment": "Neutral", "score": 0.0}
                    for _ in comments
                ]
            time.sleep(2)
        except anthropic.RateLimitError:
            wait = 10 * attempt
            print(f"  [WARN] Rate limit – waiting {wait}s …")
            time.sleep(wait)


def analyze_comments(df: pd.DataFrame, batch_size: int = 20) -> pd.DataFrame:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY environment variable not set.\n"
            "Run:  export ANTHROPIC_API_KEY='your-key-here'"
        )

    client = anthropic.Anthropic(api_key=api_key)
    comments = df["content"].tolist()
    total = len(comments)

    topics_list, keywords_list, sentiment_list, score_list = [], [], [], []

    for start in range(0, total, batch_size):
        batch = comments[start : start + batch_size]
        end = min(start + batch_size, total)
        print(f"[ANALYZE] Processing {start + 1}–{end} / {total} …")

        results = analyze_batch(client, batch)

        for r in results:
            topics_list.append(", ".join(r.get("topics", [])[:3]))
            keywords_list.append(", ".join(r.get("keywords", [])[:5]))
            sentiment_list.append(r.get("sentiment", "Neutral"))
            score = float(r.get("score", 0.0))
            score_list.append(round(max(-1.0, min(1.0, score)), 4))  # clamp to [-1, 1]

        # Small pause to stay within rate limits
        time.sleep(0.5)

    df = df.copy()
    df["topics"]          = topics_list
    df["keywords"]        = keywords_list
    df["sentiment_label"] = sentiment_list
    df["sentiment_score"] = score_list
    return df


# ─────────────────────────────────────────────
# STEP 3 – OUTPUT  (formatted Excel)
# ─────────────────────────────────────────────

SENTIMENT_COLORS = {
    "Positive": "C6EFCE",   # green
    "Neutral":  "FFEB9C",   # yellow
    "Negative": "FFC7CE",   # red
}


def save_excel(df: pd.DataFrame, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned & Analyzed"

    # ── Column order ─────────────────────────────
    cols = ["username", "timestamp", "content", "language", "likes",
            "topics", "keywords", "sentiment_label", "sentiment_score"]
    df = df[[c for c in cols if c in df.columns]]

    # ── Header style ─────────────────────────────
    header_fill = PatternFill("solid", start_color="2F5496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    ws.append(cols)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
    ws.row_dimensions[1].height = 28

    # ── Data rows ────────────────────────────────
    normal_font = Font(name="Arial", size=9)
    wrap_align  = Alignment(vertical="top", wrap_text=True)

    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in cols])
        r_idx = ws.max_row

        sentiment = str(row.get("sentiment_label", "Neutral"))
        row_fill  = PatternFill("solid", start_color=SENTIMENT_COLORS.get(sentiment, "FFFFFF"))

        for col_idx, cell in enumerate(ws[r_idx], 1):
            cell.font      = normal_font
            cell.alignment = wrap_align
            cell.border    = thin_border
            # Highlight only the sentiment columns
            if cols[col_idx - 1] in ("sentiment_label", "sentiment_score"):
                cell.fill = row_fill

    # ── Column widths ────────────────────────────
    widths = {
        "username": 18, "timestamp": 20, "content": 60,
        "language": 10, "likes": 8,  "topics": 30,
        "keywords": 35, "sentiment_label": 14, "sentiment_score": 14,
    }
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)

    # ── Summary sheet ─────────────────────────────
    ws2 = wb.create_sheet("Summary")
    counts = df["sentiment_label"].value_counts()
    total  = len(df)

    ws2.append(["Sentiment", "Count", "Percentage"])
    for cell in ws2[1]:
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.alignment = center

    for label, cnt in counts.items():
        ws2.append([label, cnt, f"=B{ws2.max_row + 1 - 1}/B{ws2.max_row + len(counts) - counts.index.tolist().index(label)}"])

    # Simple counts without formula complexity
    ws2.delete_rows(2, ws2.max_row)
    for label in ["Positive", "Neutral", "Negative"]:
        cnt = int(counts.get(label, 0))
        pct = f"{cnt / total * 100:.1f}%" if total else "0.0%"
        row_obj = ws2.append([label, cnt, pct])
        r = ws2.max_row
        fill = PatternFill("solid", start_color=SENTIMENT_COLORS.get(label, "FFFFFF"))
        for c in ws2[r]:
            c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = center

    ws2.append(["Total cleaned", total, "100.0%"])
    for c in ws2[ws2.max_row]:
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = center

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 18

    wb.save(out_path)
    print(f"\n[DONE] Saved → {out_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="YouTube comment cleaner + sentiment analyzer")
    parser.add_argument("input",  help="Path to input CSV or Excel file")
    parser.add_argument("-o", "--output", default=None,
                        help="Output Excel path (default: <input>_analyzed.xlsx)")
    parser.add_argument("--min-len",    type=int, default=40,  help="Min comment length (default 40)")
    parser.add_argument("--max-len",    type=int, default=500, help="Max comment length (default 500)")
    parser.add_argument("--batch-size", type=int, default=20,  help="Comments per API call (default 20)")
    parser.add_argument("--skip-ai",    action="store_true",
                        help="Only clean data, skip AI analysis (useful for testing)")
    args = parser.parse_args()

    # ── Load ──────────────────────────────────
    p = Path(args.input)
    if p.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(p)
    else:
        df = pd.read_csv(p)

    # Normalise column names
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    required = {"username", "timestamp", "content", "language", "likes"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    # ── Clean ─────────────────────────────────
    df_clean = clean_dataframe(df, args.min_len, args.max_len)

    if df_clean.empty:
        print("[WARN] No comments left after cleaning. Exiting.")
        return

    # ── Analyze ───────────────────────────────
    if not args.skip_ai:
        df_final = analyze_comments(df_clean, batch_size=args.batch_size)
    else:
        print("[SKIP] AI analysis skipped (--skip-ai flag).")
        df_final = df_clean

    # ── Save ──────────────────────────────────
    out = args.output or str(p.with_name(p.stem + "_analyzed.xlsx"))
    save_excel(df_final, out)

    # ── Quick stats ───────────────────────────
    if "sentiment_label" in df_final.columns:
        print("\n── Sentiment Summary ──────────────────")
        for label, cnt in df_final["sentiment_label"].value_counts().items():
            pct = cnt / len(df_final) * 100
            print(f"  {label:<10}: {cnt:>5,}  ({pct:.1f}%)")
        avg = df_final["sentiment_score"].mean()
        print(f"\n  Avg sentiment score : {avg:+.4f}")
        print("────────────────────────────────────────")


if __name__ == "__main__":
    main()
